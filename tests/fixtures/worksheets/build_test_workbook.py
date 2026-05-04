"""Generate small .xlsx fixture files for parser/worksheet tests.

Run this script when you change the fixture shape:
    python tests/fixtures/worksheets/build_test_workbook.py

It writes minimal_valid.xlsx, missing_tiers_column.xlsx, with_substitution.xlsx
into this directory, mirroring the layout of the real Riverside worksheet
(title rows, summary block, base table, enhancements table, tier scenarios).
"""
from __future__ import annotations

from pathlib import Path

import openpyxl

HERE = Path(__file__).parent

HEADER_FULL = [
    "#", "Item", "Description / Location", "Qty", "Unit",
    "Price\nper Unit", "Line Total", "Rendering Reference",
    "Materials / Build / Anchoring", "Notes / Assumptions",
    "Customer-Facing Description", "Zone", "Tiers",
]

HEADER_NO_TIERS = HEADER_FULL[:-1]   # drops the Tiers column


def _write_workbook(path: Path, header: list[str], base_rows: list[list], enh_rows: list[list],
                    scenarios: list[tuple[str, float]]) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Scope Worksheet"

    # Title rows
    ws.cell(row=1, column=1, value="Test Project — Scope Worksheet")
    ws.cell(row=2, column=1, value="Test header row")
    # Blank row 3
    # Header row 4
    for col, h in enumerate(header, start=1):
        ws.cell(row=4, column=col, value=h)
    # Base rows starting row 5
    next_row = 5
    for r in base_rows:
        for col, val in enumerate(r, start=1):
            ws.cell(row=next_row, column=col, value=val)
        next_row += 1
    # Blank row, then total row, then blank
    ws.cell(row=next_row, column=2, value="BASE SCOPE TOTAL — Year 1")
    ws.cell(row=next_row, column=7, value=sum(r[6] for r in base_rows))
    next_row += 2

    # OPTIONAL ENHANCEMENTS section header
    ws.cell(row=next_row, column=1, value="OPTIONAL ENHANCEMENTS — priced individually")
    next_row += 1
    # Enhancements header row (same shape)
    for col, h in enumerate(header, start=1):
        ws.cell(row=next_row, column=col, value=h)
    next_row += 1
    for r in enh_rows:
        for col, val in enumerate(r, start=1):
            ws.cell(row=next_row, column=col, value=val)
        next_row += 1
    next_row += 1

    # TIER SCENARIOS block
    ws.cell(row=next_row, column=1, value="TIER SCENARIOS")
    next_row += 1
    for label, total in scenarios:
        ws.cell(row=next_row, column=2, value=label)
        ws.cell(row=next_row, column=7, value=total)
        next_row += 1

    wb.save(path)


def _row(line_num, item, qty, price, customer_facing, zone, tiers):
    """Build a 13-column row. Internal description, rendering, materials are blank for tests."""
    line_total = qty * price
    return [
        line_num, item, "internal desc", qty, "ea", price, line_total,
        "rendering.png", "materials", "notes",
        customer_facing, zone, tiers,
    ]


# minimal_valid.xlsx — 2 base + 1 enhancement, all in all 3 tiers
def build_minimal_valid() -> None:
    base = [
        _row("1", "Wreath", 4, 100, "Lighted wreaths at the entrance.", "Zone One",
             "Essential, Enhanced, Signature"),
        _row("2", "Garland", 100, 25, "Lit garland on the perimeter fence.", "*",
             "Essential, Enhanced, Signature"),
    ]
    enh = [
        _row("E1", "Snowflakes", 12, 295, "Lighted snowflakes on platform railings.",
             "Zone One", "Enhanced, Signature"),
    ]
    base_total = sum(r[6] for r in base)
    enh_total = sum(r[6] for r in enh)
    scenarios = [
        ("ESSENTIAL — Base only", base_total),
        ("ENHANCED — Base + Snowflakes", base_total + enh_total),
        ("SIGNATURE — All", base_total + enh_total),
    ]
    _write_workbook(HERE / "minimal_valid.xlsx", HEADER_FULL, base, enh, scenarios)


# missing_tiers_column.xlsx — header omits Tiers column
def build_missing_tiers() -> None:
    base = [
        ["1", "Wreath", "internal", 4, "ea", 100, 400, "render.png", "mat", "notes",
         "Lit wreaths.", "*"],   # only 12 cols, no Tiers
    ]
    _write_workbook(HERE / "missing_tiers_column.xlsx", HEADER_NO_TIERS, base, [], [])


# with_substitution.xlsx — Traditional Tree (Essential, Enhanced) + Spiral LED (Signature)
def build_substitution() -> None:
    base = [
        _row("1", "Traditional Tree", 1, 18000, "Traditional centerpiece tree.",
             "Zone One", "Essential, Enhanced"),
    ]
    enh = [
        _row("E1", "Spiral LED Tree", 1, 22000, "Spiral LED replacement tree.",
             "Zone One", "Signature"),
    ]
    scenarios = [
        ("ESSENTIAL", 18000),
        ("ENHANCED", 18000),
        ("SIGNATURE", 22000),    # Traditional excluded; Spiral LED included
    ]
    _write_workbook(HERE / "with_substitution.xlsx", HEADER_FULL, base, enh, scenarios)


if __name__ == "__main__":
    build_minimal_valid()
    build_missing_tiers()
    build_substitution()
    print("Generated 3 fixture .xlsx files in", HERE)
