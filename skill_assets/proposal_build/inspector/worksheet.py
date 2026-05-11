"""Worksheet-readiness checks."""
from __future__ import annotations

import re
from pathlib import Path

import frontmatter
from openpyxl import load_workbook

from proposal_build.inspector.report import Finding


# Mirrors parser.worksheet._LINE_NUM_RE — a data row's first cell looks
# like "1", "12", or "E6" (Enhancement). Anything else (blank, summary
# row, sub-header) is skipped by the inspector to avoid false-positive
# blockers on rows that aren't supposed to carry line-item data.
_LINE_NUM_RE = re.compile(r"^(?:\d+|E\d+)$")


SCOPE_DIR = "03 - Scope & Pricing"
WORKSHEET_SUFFIX = " - Scope Worksheet.xlsx"
BRIEF_RELPATH = "04 - Process & Notes/Project Brief.md"
# Required column names mirror parser.worksheet.REQUIRED_HEADERS. Keep in sync.


def _read_project_name(project_path: Path) -> str | None:
    """Best-effort read of project_name from the Brief frontmatter.

    Returns None if the Brief is missing or unparseable — those problems
    are flagged separately by inspector/brief.py, so this helper just
    skips the name-mismatch check rather than duplicating the finding.
    """
    brief_path = project_path / BRIEF_RELPATH
    if not brief_path.is_file():
        return None
    try:
        post = frontmatter.load(str(brief_path))
        name = (post.metadata or {}).get("project_name")
        return str(name) if name else None
    except Exception:
        return None


def _find_worksheet(scope_dir: Path, expected_name: str | None = None) -> tuple[Path | None, Path | None]:
    """Locate the worksheet to inspect, plus the path the parser will look for.

    Returns (path_used, path_expected). When expected_name is given and that
    exact file exists, both are the same. When the expected file is missing
    but a glob-matching file exists, returns (glob_match, expected_path) so
    the caller can flag the name mismatch. If no expected_name is supplied
    (Brief unavailable), falls back to the original glob-match behavior.
    """
    expected_path = (scope_dir / f"{expected_name}{WORKSHEET_SUFFIX}") if expected_name else None
    if expected_path is not None and expected_path.is_file():
        return expected_path, expected_path
    glob_match: Path | None = None
    for p in scope_dir.glob(f"*{WORKSHEET_SUFFIX}"):
        if not p.name.startswith(".~lock."):
            glob_match = p
            break
    return glob_match, expected_path


def _is_locked(worksheet_path: Path) -> bool:
    return (worksheet_path.parent /
            f".~lock.{worksheet_path.name}#").exists()


def check(project_path: Path) -> list[Finding]:
    findings: list[Finding] = []
    scope_dir = project_path / SCOPE_DIR
    if not scope_dir.is_dir():
        # folder.py already reports this; don't duplicate.
        return []

    # If the Brief is menu-mode, skip the tiered-shape worksheet checks
    # and run the ROM equivalents.
    brief_path = project_path / BRIEF_RELPATH
    mode = "tiered"
    if brief_path.is_file():
        try:
            mode = frontmatter.load(str(brief_path)).metadata.get("mode", "tiered")
        except Exception:
            pass
    if mode == "menu":
        return _check_menu_worksheet(project_path)

    expected_name = _read_project_name(project_path)
    ws_path, expected_path = _find_worksheet(scope_dir, expected_name)
    if ws_path is None:
        findings.append(Finding(
            severity="blocker", category="worksheet",
            issue="missing-worksheet",
            detail=f"No `*{WORKSHEET_SUFFIX}` file found in {SCOPE_DIR}/",
            fix=("Scaffold the project (or copy the template Worksheet "
                 f"into `{SCOPE_DIR}/<Project Name>{WORKSHEET_SUFFIX}`)."),
        ))
        return findings

    # When the Brief tells us what filename to expect (project_name + suffix)
    # but on-disk has a differently-named worksheet, the generator parser will
    # fail to load it. Block at inspect time so the mismatch surfaces here,
    # not three commands later.
    if expected_path is not None and ws_path != expected_path:
        findings.append(Finding(
            severity="blocker", category="worksheet",
            issue="worksheet-name-mismatch",
            detail=(f"Worksheet on disk is `{ws_path.name}` but the generator "
                    f"will look for `{expected_path.name}` (derived from "
                    f"`project_name` in the Brief). Generation would fail."),
            fix=f"Rename the worksheet file to `{expected_path.name}`.",
        ))
        return findings

    if _is_locked(ws_path):
        findings.append(Finding(
            severity="error", category="worksheet",
            issue="worksheet-locked",
            detail=f"Worksheet appears to be open in Excel or LibreOffice: {ws_path.name}",
            fix="Close the file in Excel (or any other app that has it open) and re-run inspect.",
        ))
        return findings

    wb = None
    try:
        wb = load_workbook(ws_path, data_only=True, read_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
    except Exception as exc:
        findings.append(Finding(
            severity="error", category="worksheet",
            issue="worksheet-read-error",
            detail=f"Could not read worksheet: {exc}",
            fix="Open the worksheet in Excel and check for corruption.",
        ))
        return findings
    finally:
        if wb is not None:
            wb.close()

    if not rows:
        findings.append(Finding(
            severity="blocker", category="worksheet",
            issue="empty-worksheet",
            detail="Worksheet has no rows.",
            fix="Add header + line-item rows to the worksheet.",
        ))
        return findings

    # The header row isn't necessarily row 0 — Riverside-style worksheets
    # have a title block + pricing summary first. Search for the first row
    # containing both required column names, mirroring the parser's
    # _find_header_row strategy.
    header_row_idx = None
    for i, row in enumerate(rows):
        cells = [str(c).strip() if c is not None else "" for c in row]
        if "Customer-Facing Description" in cells and "Tiers" in cells:
            header_row_idx = i
            break

    if header_row_idx is None:
        # Neither column exists anywhere — flag both as missing.
        findings.append(Finding(
            severity="blocker", category="worksheet",
            issue="missing-customer-facing-column",
            detail="Worksheet has no `Customer-Facing Description` column.",
            fix="Restore the column from the template.",
        ))
        findings.append(Finding(
            severity="blocker", category="worksheet",
            issue="missing-tiers-column",
            detail="Worksheet has no `Tiers` column.",
            fix="Restore the column from the template.",
        ))
        return findings

    header = [str(c).strip() if c is not None else "" for c in rows[header_row_idx]]
    try:
        cf_col = header.index("Customer-Facing Description")
    except ValueError:
        cf_col = None
    try:
        tiers_col = header.index("Tiers")
    except ValueError:
        tiers_col = None

    if cf_col is None:
        findings.append(Finding(
            severity="blocker", category="worksheet",
            issue="missing-customer-facing-column",
            detail="Worksheet has no `Customer-Facing Description` column.",
            fix="Restore the column from the template.",
        ))
    if tiers_col is None:
        findings.append(Finding(
            severity="blocker", category="worksheet",
            issue="missing-tiers-column",
            detail="Worksheet has no `Tiers` column.",
            fix="Restore the column from the template.",
        ))

    if cf_col is None or tiers_col is None:
        return findings

    for i, row in enumerate(rows[header_row_idx + 1:], start=header_row_idx + 2):
        # Only check rows that look like data rows (line number "1" or "E6").
        # Skips blank rows, sub-section headers ("OPTIONAL ENHANCEMENTS"), and
        # summary rows that follow the data block.
        first_cell = str(row[0]).strip() if row and row[0] is not None else ""
        if not _LINE_NUM_RE.match(first_cell):
            continue
        line_num = first_cell
        cf_val = row[cf_col] if cf_col < len(row) else None
        tiers_val = row[tiers_col] if tiers_col < len(row) else None

        if cf_val is None or str(cf_val).strip() == "":
            findings.append(Finding(
                severity="blocker", category="worksheet",
                issue="blank-customer-facing",
                detail=f"Line {line_num} has no Customer-Facing Description.",
                fix=f"Fill in the customer-facing copy for line {line_num}.",
            ))
        if tiers_val is None or str(tiers_val).strip() == "":
            findings.append(Finding(
                severity="blocker", category="worksheet",
                issue="no-tiers-on-line",
                detail=f"Line {line_num} has no Tiers assigned.",
                fix=(f"Add a comma-separated tier list for line {line_num} "
                     "(Essential, Enhanced, Signature)."),
            ))

    return findings


def _read_project_short(project_path: Path) -> str | None:
    brief_path = project_path / BRIEF_RELPATH
    if not brief_path.is_file():
        return None
    try:
        post = frontmatter.load(str(brief_path))
        short = (post.metadata or {}).get("project_short")
        return str(short) if short else None
    except Exception:
        return None


def _check_menu_worksheet(project_path: Path) -> list[Finding]:
    """ROM worksheet readiness for menu-mode projects."""
    from proposal_build.parser.worksheet_rom import (
        parse_rom_worksheet, ROMWorksheetParseError,
    )
    from proposal_build.parser.brief import parse_brief, BriefParseError

    findings: list[Finding] = []
    scope_dir = project_path / SCOPE_DIR
    if not scope_dir.is_dir():
        findings.append(Finding(
            severity="blocker", category="worksheet", issue="no-scope-dir",
            detail=f"Missing `{SCOPE_DIR}/` folder.",
            fix=f"Create `{SCOPE_DIR}/` and place the ROM worksheet there.",
        ))
        return findings

    # Locate the worksheet (mirror the parser's _find_menu_worksheet logic).
    project_short = _read_project_short(project_path) or _read_project_name(project_path) or ""
    candidate = scope_dir / f"{project_short}{WORKSHEET_SUFFIX}"
    if candidate.is_file():
        worksheet_path = candidate
    else:
        matches = list(scope_dir.glob("*Scope Worksheet*.xlsx"))
        if not matches:
            matches = [p for p in scope_dir.glob("*.xlsx") if not p.name.startswith(".~lock.")]
        if len(matches) == 1:
            worksheet_path = matches[0]
        elif not matches:
            findings.append(Finding(
                severity="blocker", category="worksheet", issue="missing-worksheet",
                detail=f"No `.xlsx` worksheet found in `{SCOPE_DIR}/`.",
                fix=f"Place the ROM Scope Worksheet in `{SCOPE_DIR}/`.",
            ))
            return findings
        else:
            findings.append(Finding(
                severity="blocker", category="worksheet", issue="ambiguous-worksheet",
                detail=f"Multiple worksheets in `{SCOPE_DIR}/`: {[p.name for p in matches]}",
                fix="Keep only one ROM worksheet in the scope directory.",
            ))
            return findings

    # Try to parse the ROM worksheet — surfaces header/shape errors clearly.
    try:
        rom = parse_rom_worksheet(worksheet_path)
    except ROMWorksheetParseError as exc:
        findings.append(Finding(
            severity="blocker", category="worksheet", issue="rom-worksheet-parse-error",
            detail=f"Could not parse ROM worksheet {worksheet_path.name}: {exc}",
            fix="Verify the ROM worksheet has the expected 15-column header row.",
        ))
        return findings

    # Cross-reference: every item code in the Brief's sections must exist
    # in the ROM worksheet.
    brief_path = project_path / BRIEF_RELPATH
    if brief_path.is_file():
        try:
            brief = parse_brief(brief_path)
            ws_codes = {it.code for it in rom.line_items}
            for s in brief.frontmatter.get("sections", []) or []:
                for code in s.get("item_codes", []):
                    if code not in ws_codes:
                        findings.append(Finding(
                            severity="blocker", category="worksheet",
                            issue="brief-references-missing-item-code",
                            detail=(
                                f"Section {s.get('key', '?')!r} references "
                                f"item code {code!r}, which is not in the ROM worksheet."
                            ),
                            fix=(f"Add item `{code}` to the worksheet or remove it "
                                 "from the Brief's item_codes."),
                        ))
        except BriefParseError:
            pass  # brief inspector already reported this

    return findings
