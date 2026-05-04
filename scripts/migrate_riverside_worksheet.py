"""One-shot migration of Riverside MetroLink - Scope Worksheet.xlsx.

Two operations in a single pass:

1. Adds Customer-Facing Description, Zone, Tiers columns to all 25 line items.
   Zone names match the 8-zone structure in Project Brief.md.

2. Restructures the TIER SCENARIOS block + tier-model legend bullet to reflect
   the consistent-garland-treatment principle: garland decoration is uniform
   within each tier (never mixed undecorated/decorated within the same tier).
   - ESSENTIAL: $88,906 — base only, all garland undecorated
   - ENHANCED:  $124,292 — base + 5 visual-interest add-ons (E1, E2, E7, E11, E12); garland still all undecorated
   - SIGNATURE: $200,249 — enhanced + Bell Display + Spiral LED Tree (replaces
                          Traditional) + ALL four garland sections decorated

Idempotent: running twice produces the same result. Kept in repo for
reproducibility — future projects can model after it.
"""
from __future__ import annotations

from pathlib import Path

import openpyxl

REPO = Path(__file__).resolve().parent.parent
WORKSHEET = REPO / "Projects" / "Downtown Riverside Metro Link" / "03 - Scope & Pricing" / "Riverside MetroLink - Scope Worksheet.xlsx"

# (line_num, customer_facing, zone, tiers)
# Customer-facing copy follows: Size → Item name → Color/material/details.
RIVERSIDE_MIGRATION = {
    "1":  ("24″×48″ branded 'Holiday Express' pole banners — 18 oz vinyl, double-sided, hemmed + grommeted (customer-owned)",
           "Pole Banner Program", "Essential, Enhanced, Signature"),
    "2":  ("Heavy-duty powder-coated steel pole banner brackets — top + bottom arm pair, stainless banding mount (no drilling, customer-owned)",
           "Pole Banner Program", "Essential, Enhanced, Signature"),
    "3":  ("Annual install + removal of 35 pole banners — bucket-truck access, off-season storage included",
           "Pole Banner Program", "Essential, Enhanced, Signature"),
    "4":  ("Commercial-grade C9 LED warm-white string lighting on 16 platform canopies — 1,024 LF on the eaves",
           "Canopy & Platform Lighting", "Essential, Enhanced, Signature"),
    "5":  ("Commercial-grade C9 LED warm-white string lighting on 2 bus-stop waiting canopies — 234 LF on the eaves",
           "Canopy & Platform Lighting", "Essential, Enhanced, Signature"),
    "6":  ("14″ PVC pine garland with warm-white mini-LED lights — 621 LF on the perimeter fence",
           "Perimeter & Driveway Garlands", "Essential, Enhanced, Signature"),
    "7":  ("14″ PVC pine garland with warm-white mini-LED lights — 144 LF on the building eave",
           "Perimeter & Driveway Garlands", "Essential, Enhanced, Signature"),
    "8":  ("14″ PVC pine garland with warm-white mini-LED lights — 261 LF on the center driveway gates",
           "Perimeter & Driveway Garlands", "Essential, Enhanced, Signature"),
    "9":  ("5 ft pre-lit warm-white LED commercial wreaths with red, green & gold ornament clusters — 4 between the entrance brick columns",
           "Station Entrance Wreaths", "Essential, Enhanced, Signature"),
    "10": ("10 ft oversized pre-lit warm-white LED commercial wreath with red, green & gold ornament clusters — mounted on the stair-tower brick column",
           "Stair Tower Feature", "Essential, Enhanced, Signature"),
    "11": ("18 ft pre-lit artificial steel-frame Christmas tree — warm-white LED with red, green & gold ornaments, 2 ft LED star topper",
           "Plaza Centerpiece", "Essential, Enhanced"),
    "12": ("12 ft warm-white LED walk-through ornament archway with companion ornament — steel armature, ground-anchored at bus loading island",
           "Walk-Through Photo Moments", "Essential, Enhanced, Signature"),
    "E1": ("30″ aluminum-frame warm-white rope-light snowflakes — 12 fixtures railing-mounted along bridge and platforms",
           "Signature Add-Ons", "Enhanced, Signature"),
    "E2": ("12 ft lighted walk-through gift box archway with red bow — steel frame, LED + PVC mesh, ground-anchored",
           "Walk-Through Photo Moments", "Enhanced, Signature"),
    "E3": ("10 ft custom-fabricated Mission Inn-style 'City of Riverside' lighted bell display — steel armature, warm-white LED + rope light, PVC mesh, branded acrylic base (customer-owned)",
           "Signature Add-Ons", "Signature"),
    "E4": ("Annual install + removal of the City of Riverside Bell Display — positioning, anchoring, electrical tie-in",
           "Signature Add-Ons", "Signature"),
    "E5": ("Off-season climate-controlled crated storage of the City of Riverside Bell Display",
           "Signature Add-Ons", "Signature"),
    "E6": ("21 ft steel-frame red-and-green LED spiral tree with gold star topper — alternating LED spiral wraps, warm-white rope light, PVC mesh (Signature alt. to Traditional centerpiece)",
           "Plaza Centerpiece", "Signature"),
    "E7": ("12 ft stacked lighted gift box tower — multiple aluminum-frame boxes with red, green & white LED net wrap, ground-anchored at second plaza location",
           "Signature Add-Ons", "Enhanced, Signature"),
    "E8":  ("Decorated upgrade — red, green & gold ornament clusters and bows on the 621 LF perimeter fence garland",
            "Perimeter & Driveway Garlands", "Signature"),
    "E9":  ("Decorated upgrade — red, green & gold ornament clusters and bows on the 144 LF building eave garland",
            "Perimeter & Driveway Garlands", "Signature"),
    "E10": ("Decorated upgrade — red, green & gold ornament clusters and bows on the 261 LF center driveway garland",
            "Perimeter & Driveway Garlands", "Signature"),
    "E11": ("12 ft wide custom-fabricated illuminated 'Happy Holidays' overhead marquee with skyline silhouette — PVC mesh, warm-white channel-letter LED + rope light, span-mounted between columns",
            "Plaza Centerpiece", "Enhanced, Signature"),
    "E12": ("14″ PVC pine garland with warm-white mini-LED lights — 207 LF on the staircase tower and railing",
            "Stair Tower Feature", "Enhanced, Signature"),
    "E13": ("Decorated upgrade — red, green & gold ornament clusters and bows on the 207 LF staircase tower garland",
            "Stair Tower Feature", "Signature"),
}

# Tier-scenarios block + legend updates (operation 2 — see module docstring).
# Row positions: R43 = Enhanced; R44 = Signature; R57 = tier-model legend bullet.
TIER_SCENARIOS = {
    43: (
        "ENHANCED ⭐ — Base + Snowflakes + Walk-Through Gift Box + Gift Box Tower + Happy Holidays Sign + Staircase Garland (Undecorated)",
        124292,
    ),
    44: (
        "SIGNATURE — Enhanced + Bell Display + Spiral LED Tree (replaces Traditional) + All Garlands Decorated",
        200249,
    ),
}
TIER_MODEL_BULLET = (
    "• Tier model: garland decoration treatment is uniform within each tier — never mixed. "
    "ESSENTIAL = base scope, all garland undecorated. "
    "ENHANCED = base + Snowflakes + Walk-Through Gift Box + Gift Box Tower (new location) + Happy Holidays Sign + Staircase Garland (still undecorated). "
    "SIGNATURE = Enhanced + Bell Display + Spiral LED Tree (replaces Traditional) + ALL four garland sections upgraded to decorated."
)
TIER_MODEL_BULLET_ROW = 57


def main() -> None:
    if not WORKSHEET.exists():
        raise SystemExit(f"Worksheet not found: {WORKSHEET}")

    # openpyxl strips formula caches on save, breaking parser's data_only reads.
    # Workaround: snapshot all cached values up front, then write them back as
    # static values for any cell that currently holds a formula. Tradeoff —
    # the worksheet's formulas (Line Total, TIER SCENARIOS sums, etc.) become
    # hardcoded numbers. Editing Qty/Price/Unit in Excel afterwards will not
    # auto-update Line Total; AE must update Line Total manually if quantities
    # change. Customer-facing description edits (the polish workflow) are
    # unaffected.
    wb_values = openpyxl.load_workbook(str(WORKSHEET), data_only=True)
    ws_values = wb_values.active
    cached: dict[tuple[int, int], object] = {}
    for row in ws_values.iter_rows():
        for cell in row:
            if cell.value is not None:
                cached[(cell.row, cell.column)] = cell.value

    wb = openpyxl.load_workbook(str(WORKSHEET))
    ws = wb.active

    # Convert all formula cells to their cached static values
    formulas_replaced = 0
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith("="):
                cached_val = cached.get((cell.row, cell.column))
                if cached_val is not None:
                    cell.value = cached_val
                    formulas_replaced += 1
    print(f"Converted {formulas_replaced} formulas to static values.")

    # Find header row(s). The Riverside file has TWO header rows
    # (one for Base Scope, one for Enhancements) — both must get the new columns.
    rows = list(ws.iter_rows(values_only=False))
    header_indices = []
    for i, row in enumerate(rows):
        if row[0].value == "#" and row[1].value == "Item":
            header_indices.append(i + 1)   # openpyxl is 1-indexed
    print(f"Found {len(header_indices)} header rows at: {header_indices}")

    # Add the 3 new column headers after the existing 10
    NEW_HEADERS = ("Customer-Facing Description", "Zone", "Tiers")
    for hi in header_indices:
        for offset, name in enumerate(NEW_HEADERS, start=11):
            ws.cell(row=hi, column=offset, value=name)

    # Walk all data rows; fill in the 3 new columns from RIVERSIDE_MIGRATION
    filled = 0
    for row in ws.iter_rows():
        line_num = row[0].value
        if line_num is None:
            continue
        line_str = str(line_num).strip()
        if line_str in RIVERSIDE_MIGRATION:
            cf, zone, tiers = RIVERSIDE_MIGRATION[line_str]
            row[10].value = cf
            row[11].value = zone
            row[12].value = tiers
            filled += 1
    print(f"Filled {filled} data rows with the 3 new columns.")

    # Operation 2: tier-scenarios block + tier-model legend bullet
    for row_idx, (label, total) in TIER_SCENARIOS.items():
        ws.cell(row=row_idx, column=2, value=label)
        ws.cell(row=row_idx, column=7, value=total)
    ws.cell(row=TIER_MODEL_BULLET_ROW, column=1, value=TIER_MODEL_BULLET)
    print(f"Updated {len(TIER_SCENARIOS)} tier-scenarios rows + 1 legend bullet.")

    wb.save(str(WORKSHEET))
    print(f"Saved: {WORKSHEET}")


if __name__ == "__main__":
    main()
