"""Render the FIGat7th DTLA customer-facing scope & pricing workbook.

Per feedback_customer_facing_workbook.md: every project ships a clean,
brand-styled xlsx alongside the proposal PDF + itemized pricing PDFs.
Drops internal columns (Materials/Build, Notes, Rendering Ref, Item codes)
and presents the deck-mode sections with ROM rental + purchase columns
the customer can pick across.

Usage:
    python scripts/render_figat7th_customer_workbook.py
"""
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from proposal_build.parser import parse_project
from proposal_build.composer.menu_ctx_builders import build_menu_rom_investment_ctx


REPO_ROOT = Path(__file__).resolve().parents[1]
PROJ = REPO_ROOT / "Projects" / "Fig at 7th - 2026 - Multi-Rendering Project"
OUT = PROJ / "03 - Scope & Pricing" / "FIGat7th DTLA - 2026 Holiday Scope & Pricing.xlsx"


# Brand colors (from skill_assets/layouts/brand.css), without the leading #
RED = "B31315"
CHARCOAL = "1C1C1C"
GRAY = "555555"
LIGHT = "ECEFF1"
PANEL = "F2F2F2"
WARM_TINT = "FFF6F0"     # very pale red wash for section bands
WHITE = "FFFFFF"

THIN = Side(border_style="thin", color="D4D4D4")
RULE = Side(border_style="thin", color=CHARCOAL)


def _fmt_range(low: int, high: int) -> str:
    if low == high:
        return f"${low:,}"
    return f"${low:,} – ${high:,}"


def _section_caveat(key: str) -> str:
    if key == "3a":
        return " · Customer picks one (alternate group)"
    if key == "3b":
        return " · All items included"
    return ""


def main() -> None:
    model = parse_project(PROJ)
    rom = build_menu_rom_investment_ctx(model, page_num=1, page_total=1, page_part=2)

    wb = Workbook()
    ws = wb.active
    ws.title = "Scope & Pricing"

    # Column widths tuned for landscape print-fit
    widths = {
        "A": 24,  # Section / Item
        "B": 56,  # Customer-facing description
        "C": 18,  # Rental
        "D": 22,  # Purchase (one-time)
        "E": 22,  # Service (annual)
    }
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    row = 1

    # === Title block ===
    ws.cell(row=row, column=1, value=f"{model.client_company} — Holiday Scope & Pricing").font = Font(
        name="Calibri", size=22, bold=True, color=CHARCOAL,
    )
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    ws.row_dimensions[row].height = 30
    row += 1

    ws.cell(row=row, column=1, value=f"{model.project_year} {model.project_subtitle or 'Holiday Proposal'}").font = Font(
        name="Calibri", size=12, bold=True, color=RED,
    )
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    row += 1

    ws.cell(
        row=row, column=1,
        value=f"Prepared by St. Nick's Christmas Lighting & Décor  ·  ST-NICKS.COM  ·  (562) 438-0017",
    ).font = Font(name="Calibri", size=10, italic=True, color=GRAY)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    row += 1

    # spacer
    row += 1

    # === Note row ===
    note = (
        "All items are priced à la carte. Each item can be priced as Rental "
        "(single all-inclusive annual fee covering item + install + removal + storage) "
        "or Purchase (one-time cost plus an annual Service fee). "
        "ROM ranges are budgetary; final fixed pricing is set when scope is locked."
    )
    note_cell = ws.cell(row=row, column=1, value=note)
    note_cell.font = Font(name="Calibri", size=10, color=CHARCOAL)
    note_cell.alignment = Alignment(wrap_text=True, vertical="center")
    note_cell.fill = PatternFill("solid", fgColor=PANEL)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    ws.row_dimensions[row].height = 48
    row += 1
    row += 1  # spacer

    # === Header ===
    header_row = row
    headers = ["Section / Item", "Customer-Facing Description",
               "Rental (annual)", "Purchase (one-time)", "Service (annual)"]
    for col_idx, label in enumerate(headers, start=1):
        c = ws.cell(row=row, column=col_idx, value=label)
        c.font = Font(name="Calibri", size=10, bold=True, color=LIGHT)
        c.fill = PatternFill("solid", fgColor=CHARCOAL)
        c.alignment = Alignment(horizontal="left" if col_idx <= 2 else "right",
                                vertical="center", wrap_text=True)
        c.border = Border(top=RULE, bottom=RULE, left=THIN, right=THIN)
    ws.row_dimensions[row].height = 28
    row += 1

    # === Section + item rows ===
    for section in model.sections:
        # Section band
        band = ws.cell(row=row, column=1,
                       value=f"{section.label}{_section_caveat(section.key)}")
        band.font = Font(name="Calibri", size=11, bold=True, color=CHARCOAL)
        band.fill = PatternFill("solid", fgColor=WARM_TINT)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        for col_idx in range(1, 6):
            ws.cell(row=row, column=col_idx).border = Border(top=THIN, bottom=THIN)
            ws.cell(row=row, column=col_idx).fill = PatternFill("solid", fgColor=WARM_TINT)
        ws.row_dimensions[row].height = 22
        row += 1

        # Item rows
        for li in section.items:
            ws.cell(row=row, column=1, value=li.name).font = Font(
                name="Calibri", size=10, bold=True, color=CHARCOAL,
            )
            desc = ws.cell(row=row, column=2, value=li.customer_facing or li.description)
            desc.font = Font(name="Calibri", size=10, color=CHARCOAL)
            desc.alignment = Alignment(wrap_text=True, vertical="top")

            for col_idx, val in enumerate((
                _fmt_range(li.rental_low, li.rental_high),
                _fmt_range(li.purchase_ot_low, li.purchase_ot_high),
                _fmt_range(li.purchase_svc_low, li.purchase_svc_high),
            ), start=3):
                c = ws.cell(row=row, column=col_idx, value=val)
                c.font = Font(name="Calibri", size=10, color=CHARCOAL)
                c.alignment = Alignment(horizontal="right", vertical="top")

            for col_idx in range(1, 6):
                ws.cell(row=row, column=col_idx).border = Border(bottom=THIN)
            ws.row_dimensions[row].height = 38
            row += 1

    # === Totals row ===
    row += 1
    totals_label = ws.cell(row=row, column=1, value="PROGRAM ROM TOTAL")
    totals_label.font = Font(name="Calibri", size=11, bold=True, color=RED)
    totals_label.alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)

    for col_idx, val in enumerate((rom["total_rental"], rom["total_purchase_ot"], rom["total_purchase_svc"]), start=3):
        c = ws.cell(row=row, column=col_idx, value=val)
        c.font = Font(name="Calibri", size=11, bold=True, color=RED)
        c.alignment = Alignment(horizontal="right", vertical="center")
    for col_idx in range(1, 6):
        ws.cell(row=row, column=col_idx).border = Border(top=RULE, bottom=RULE)
    ws.row_dimensions[row].height = 26
    row += 1
    row += 1

    # === Footnotes ===
    footnotes = [
        "Section 3a (Plaza Arches) shows four alternates; the customer picks one. Rental/Purchase totals span the lowest- to highest-priced arch option.",
        "Rental and Purchase pricing are mutually exclusive per item — the customer may mix the two modes across items but not for a single item.",
        f"Rough-Order-of-Magnitude (ROM) ranges are valid 60 days from {model.proposal_date}. Final fixed pricing is set when scope is locked.",
    ]
    for fn in footnotes:
        c = ws.cell(row=row, column=1, value=f"·  {fn}")
        c.font = Font(name="Calibri", size=9, italic=True, color=GRAY)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        ws.row_dimensions[row].height = 26
        row += 1

    # === Print + freeze ===
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)
    ws.print_options.horizontalCentered = True
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_title_rows = f"{header_row}:{header_row}"

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"Wrote: {OUT}")
    print(f"Size: {OUT.stat().st_size:,} bytes")


if __name__ == "__main__":
    main()
